"""Excel export + Crosstabs association measures."""
import numpy as np
import pandas as pd
from scipy import stats as sps

from sidecar.data.dataset import Dataset, DatasetRegistry
from sidecar.data.format import Format
from sidecar.data.variable import VariableMeta
from sidecar.procedures.registry import build_registry
from sidecar.syntax.registry import Context, execute_syntax
from sidecar.output.excel import export_xlsx


def make(df):
    reg = DatasetRegistry()
    reg.add(Dataset(df.copy(), [VariableMeta(name=c, print_format=Format("F", 8, 2)) for c in df.columns]))
    return reg


def run(reg, text):
    return execute_syntax(text, build_registry(), Context(reg))


def cells(t):
    return {(tuple(c["r"]), tuple(c["c"])): c["v"] for c in t["cells"]}


def tables(out):
    return [o for o in out if o["type"] == "PivotTable"]


def test_crosstabs_phi_cramer():
    r = [1.0] * 40 + [2.0] * 40
    c = ([1.0] * 30 + [2.0] * 10) + ([1.0] * 10 + [2.0] * 30)
    reg = make(pd.DataFrame({"r": r, "c": c}))
    out = run(reg, "CROSSTABS /TABLES=r BY c /STATISTICS=CHISQ PHI.")
    sm = [t for t in tables(out) if t["title"] == "Symmetric Measures"][0]
    obs = pd.crosstab(pd.Series(r), pd.Series(c)).to_numpy()
    chi2 = sps.chi2_contingency(obs, correction=False)[0]
    phi = (chi2 / obs.sum()) ** 0.5
    assert abs(float(cells(sm)[((0,), (0,))]) - phi) < 0.01


def test_crosstabs_gamma():
    reg = make(pd.DataFrame({"r": [1.0, 1, 2, 2, 3, 3], "c": [1.0, 2, 2, 3, 3, 3]}))
    out = run(reg, "CROSSTABS /TABLES=r BY c /STATISTICS=GAMMA.")
    sm = [t for t in tables(out) if t["title"] == "Symmetric Measures"][0]
    assert any("Gamma" in cat for cat in sm["rowDims"][0]["categories"])


def test_excel_export(tmp_path):
    reg = make(pd.DataFrame({"x": [1.0, 2, 3, 4, 5]}))
    out = run(reg, "DESCRIPTIVES VARIABLES=x.")
    p = str(tmp_path / "out.xlsx")
    export_xlsx(out, p)
    from openpyxl import load_workbook

    wb = load_workbook(p)
    ws = wb.active
    txt = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    assert "Descriptive Statistics" in txt
