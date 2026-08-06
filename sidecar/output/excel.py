"""Export an output document (list of output objects) to .xlsx via openpyxl.

Fulfils the HLD goal of exporting output to HTML, PDF, and .xlsx.
"""
from __future__ import annotations

from itertools import product
from typing import Any


def _leaf_cols(t: dict[str, Any]) -> tuple[list[str], list[tuple]]:
    if t.get("colLeaves") is not None:
        leaves = t["colLeaves"]
        return list(leaves), [(i,) for i in range(len(leaves))]
    dims = t.get("colDims", [])
    if not dims:
        return [""], [()]
    ranges = [range(len(d["categories"])) for d in dims]
    keys = list(product(*ranges))
    labels = [" / ".join(dims[k]["categories"][idx[k]] for k in range(len(dims))) for idx in keys]
    return labels, keys


def _row_tuples(t: dict[str, Any]) -> tuple[list[str], list[tuple]]:
    dims = t.get("rowDims", [])
    if not dims:
        return [""], [()]
    ranges = [range(len(d["categories"])) for d in dims]
    keys = list(product(*ranges))
    labels = [" / ".join(dims[k]["categories"][idx[k]] for k in range(len(dims))) for idx in keys]
    return labels, keys


def flatten_table(t: dict[str, Any]) -> list[list[str]]:
    cellmap = {(tuple(c["r"]), tuple(c["c"])): c["v"] for c in t["cells"]}
    col_labels, col_keys = _leaf_cols(t)
    row_labels, row_keys = _row_tuples(t)
    rows: list[list[str]] = []
    rows.append([t.get("corner", "") or ""] + col_labels)
    for rlab, rkey in zip(row_labels, row_keys):
        line = [rlab]
        for ckey in col_keys:
            line.append(cellmap.get((rkey, ckey), ""))
        rows.append(line)
    return rows


def export_xlsx(items: list[dict[str, Any]], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Output"
    r = 1
    bold = Font(bold=True)
    for obj in items:
        typ = obj.get("type")
        if typ in ("Title", "TextBlock", "Warning", "Error"):
            ws.cell(row=r, column=1, value=obj.get("text", "")).font = bold if typ == "Title" else Font()
            r += 2
        elif typ == "PivotTable":
            ws.cell(row=r, column=1, value=obj.get("title", "")).font = bold
            r += 1
            for line in flatten_table(obj):
                for c, val in enumerate(line, start=1):
                    ws.cell(row=r, column=c, value=_coerce(val))
                r += 1
            r += 1
        elif typ == "Chart":
            ws.cell(row=r, column=1, value="[Chart]")
            r += 2
    wb.save(path)


def _coerce(v: str) -> Any:
    if v in ("", ".", None):
        return v
    try:
        return float(v.replace(",", "")) if any(ch.isdigit() for ch in v) else v
    except (ValueError, AttributeError):
        return v
